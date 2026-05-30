"""Ingestion channels beyond manual upload.

  - POST /intake/email : email webhook payload  → one case
  - POST /intake/csv   : CSV/XLSX upload        → one case per data row
  - POST /intake/erp   : structured ERP JSON    → one case

All channels converge on apps.api.services.ingestion.ingest_document so they share
hash dedup, storage, and the Document+Case+AuditEvent transaction. Document/email
content is untrusted and is sanitised before it is ever stored or shown to the LLM
. Each non-duplicate case is enqueued on the shared Redis pool.
"""

from __future__ import annotations

import csv
import io
import logging
import re

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from packages.domain.enums import ActorType
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from apps.api.auth import CurrentUser, get_current_user
from apps.api.database import get_session
from apps.api.services.ingestion import (
    IngestResult,
    canonical_text_from_mapping,
    ingest_document,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/intake", tags=["intake"])

_MAX_BODY_LEN = 8_000
_MAX_CSV_ROWS = 500
# Strip anything that looks like a prompt-injection attempt.
_INJECTION_RE = re.compile(
    r"\b(ignore (all )?(previous|prior|above) instructions?|"
    r"system prompt|forget (everything|all)|you are now)\b",
    re.IGNORECASE,
)


def _sanitise(text: str) -> str:
    cleaned = _INJECTION_RE.sub("[REDACTED]", text)
    return cleaned[:_MAX_BODY_LEN]


async def _enqueue(case_id: str) -> None:
    """Best-effort enqueue on the shared pool; case still exists if Redis is down."""
    try:
        from apps.api.redis_pool import get_redis_pool

        await get_redis_pool().enqueue_job("process_document", case_id)
    except Exception:
        logger.warning("intake.enqueue_failed case_id=%s — admin can reprocess", case_id)


# ── Email ─────────────────────────────────────────────────────────────────────


class EmailIntakeRequest(BaseModel):
    from_address: str
    subject: str
    body_text: str
    attachments: list[str] = []  # filenames; actual files via separate upload


class EmailIntakeResponse(BaseModel):
    case_id: str
    document_id: str
    message: str


@router.post("/email", response_model=EmailIntakeResponse, status_code=201)
async def intake_email(
    body: EmailIntakeRequest,
    session: AsyncSession = Depends(get_session),
    user: CurrentUser = Depends(get_current_user),
) -> EmailIntakeResponse:
    """Create a Document + Case from an incoming email payload and queue it."""
    if not body.subject.strip():
        raise HTTPException(status_code=422, detail="subject must not be empty.")

    # Sanitise BEFORE storing — the stored bytes feed OCR/extraction (untrusted data).
    sanitised_body = _sanitise(body.body_text)
    document_text = f"Subject: {body.subject}\nFrom: {body.from_address}\n\n{sanitised_body}"

    result = await ingest_document(
        session,
        user.org_id,
        filename=f"email:{body.subject[:120]}",
        content_type="text/plain",
        content=document_text.encode("utf-8"),
        channel="email",
        actor_type=ActorType.HUMAN,  # an email arrived because a human sent it
        actor_id=user.user_id,
        extra_payload={"from": body.from_address, "subject": body.subject},
    )
    if not result.is_duplicate:
        await _enqueue(result.case_id)

    return EmailIntakeResponse(
        case_id=result.case_id,
        document_id=result.document_id,
        message="Email intake accepted. Case created and queued for processing.",
    )


# ── CSV / XLSX ─────────────────────────────────────────────────────────────────


class CsvIntakeResponse(BaseModel):
    case_ids: list[str]
    rows_ingested: int
    duplicates: int
    message: str


def _parse_rows(filename: str, content_type: str, raw: bytes) -> list[dict[str, object]]:
    """Parse a CSV or XLSX upload into a list of row dicts keyed by header."""
    name = (filename or "").lower()
    is_xlsx = name.endswith((".xlsx", ".xlsm")) or "spreadsheet" in (content_type or "")
    if is_xlsx:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency guard
            raise HTTPException(
                status_code=422,
                detail="XLSX support requires openpyxl. Install it or upload CSV.",
            ) from exc
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [
                str(h).strip() if h is not None else f"col{i}"
                for i, h in enumerate(next(rows_iter))
            ]
        except StopIteration:
            return []
        out: list[dict[str, object]] = []
        for values in rows_iter:
            if values is None or all(v is None for v in values):
                continue
            out.append({header[i]: values[i] for i in range(min(len(header), len(values)))})
        return out

    # CSV (stdlib): decode and let DictReader key by header.
    text = raw.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


@router.post("/csv", response_model=CsvIntakeResponse, status_code=201)
async def intake_csv(
    file: UploadFile,
    session: AsyncSession = Depends(get_session),
    user: CurrentUser = Depends(get_current_user),
) -> CsvIntakeResponse:
    """Import a CSV/XLSX of documents — one traceable case per data row."""
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    rows = _parse_rows(file.filename or "upload.csv", file.content_type or "", raw)
    if not rows:
        raise HTTPException(status_code=422, detail="No data rows found in the file.")
    if len(rows) > _MAX_CSV_ROWS:
        raise HTTPException(
            status_code=422, detail=f"Too many rows ({len(rows)} > {_MAX_CSV_ROWS})."
        )

    case_ids: list[str] = []
    duplicates = 0
    for n, row in enumerate(rows):
        document_text = _sanitise(canonical_text_from_mapping(row))
        if not document_text.strip():
            continue
        result: IngestResult = await ingest_document(
            session,
            user.org_id,
            filename=f"{file.filename or 'rows'}#row{n + 1}",
            content_type="text/plain",
            content=document_text.encode("utf-8"),
            channel="csv",
            actor_id=user.user_id,
        )
        if result.is_duplicate:
            duplicates += 1
            continue
        case_ids.append(result.case_id)
        await _enqueue(result.case_id)

    return CsvIntakeResponse(
        case_ids=case_ids,
        rows_ingested=len(case_ids),
        duplicates=duplicates,
        message=f"Ingested {len(case_ids)} case(s) from {len(rows)} row(s).",
    )


# ── ERP / API ──────────────────────────────────────────────────────────────────


class ErpIntakeRequest(BaseModel):
    # Free-form structured record from an ERP/automation. Keys are mapped to the
    # canonical document fields by canonical_text_from_mapping.
    fields: dict[str, object]
    source_system: str | None = None
    external_id: str | None = None


class ErpIntakeResponse(BaseModel):
    case_id: str
    document_id: str
    is_duplicate: bool
    message: str


@router.post("/erp", response_model=ErpIntakeResponse, status_code=201)
async def intake_erp(
    body: ErpIntakeRequest,
    session: AsyncSession = Depends(get_session),
    user: CurrentUser = Depends(get_current_user),
) -> ErpIntakeResponse:
    """Ingest a structured ERP record and run it through the standard pipeline."""
    if not body.fields:
        raise HTTPException(status_code=422, detail="fields must not be empty.")

    document_text = _sanitise(canonical_text_from_mapping(body.fields))
    if not document_text.strip():
        raise HTTPException(status_code=422, detail="No recognisable fields in payload.")

    label = body.external_id or body.source_system or "record"
    result = await ingest_document(
        session,
        user.org_id,
        filename=f"erp:{label}",
        content_type="text/plain",
        content=document_text.encode("utf-8"),
        channel="api",
        actor_id=user.user_id,
        extra_payload={"source_system": body.source_system, "external_id": body.external_id},
    )
    if not result.is_duplicate:
        await _enqueue(result.case_id)

    return ErpIntakeResponse(
        case_id=result.case_id,
        document_id=result.document_id,
        is_duplicate=result.is_duplicate,
        message="ERP record accepted and queued for processing.",
    )
